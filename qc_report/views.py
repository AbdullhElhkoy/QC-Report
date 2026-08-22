from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.core.paginator import Paginator
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.dateparse import parse_date

from .aggregations import full_daily_report, full_shift_report
from .pdf_charts import generate_report_charts
from .forms import (
    BulkLogForm,
    DCPASRowForm,
    DCPBBRowForm,
    DCPReasonLineForm,
    DCPTestsForm,
    DCPSBRowForm,
    PackingTableForm,
    GCCSummaryForm,
    build_dcp_reason_line_form,
    gcc_color_form,
    GCCColorRowForm,
    LoadingLineItemForm,
    ShiftEntryForm,
    SOPLineItemForm,
)
from .models import (
    BulkLog,
    DCPASRow,
    DCPBBRow,
    DCPReason,
    DCPReasonCategory,
    DCPReasonLine,
    DCPTests,
    DCPSBRow,
    GCCColorRow,
    GCCColor,
    GCCSummary,
    LoadingLineItem,
    LoadingProductType,
    PackingFactory,
    PackingLineItem,
    PackingType,
    PackageType,
    ProductType,
    EntryRevision,
    Shift,
    ShiftEntry,
    SOPLineItem,
    SOP_BULK_UNITS,
    SOP_PRODUCT_TYPES,
    Unit,
)

ZERO = Decimal("0")

# أسماء عربية لجداول بيانات كل وحدة (تظهر في صفحة التفاصيل وأوراق Excel)
MODEL_TITLES = {
    "SOPLineItem": "بنود SOP",
    "DCPSummary": "ملخص DCP",
    "DCPColorGrading": "تصنيف الألوان",
    "DCPUnderTest": "العينات تحت الاختبار",
    "DCPWhiteQuality": "جودة الأبيض",
    "DCPRework": "التدوير",
    "PALineItem": "بنود PA",
    "SALineItem": "بنود SA",
    "PackingLineItem": "بنود التعبئة (PA/SA)",    "GCCColorRow": "جدول GCC - الألوان",
    "GCCSummary": "SB / NC",
    "LoadingLineItem": "بنود التحميل",
    "BulkLog": "عربيات BULK",
    "DCPBBRow": "جدول B.B",
    "DCPSBRow": "جدول S.B",
    "DCPASRow": "جدول S.B AS",
    "DCPTests": "الاختبارات",
    "DCPReasonLine": "أسباب الأبيض والأحمر",
}

# حقول كل جدول بيتم التقاطها في نسخة سجل التعديل
_SNAPSHOT_RELATED = [
    ("sop_line_items", [
        "product_type", "exp", "dom", "std", "nc", "cause", "note",
        "defect_reason", "notes", "car_number", "car_weight",
    ]),
    ("bulk_log", ["exp_trucks", "dom_trucks", "std_trucks", "nc_trucks"]),
    ("dcp_bb", ["green", "yellow", "green_yellow", "blue", "white", "red", "note"]),
    ("dcp_sb", ["exp", "dom", "sb_white", "note"]),
    ("dcp_as", ["exp", "dom", "sb_white", "note"]),
    ("dcp_tests", ["lab_test", "floor_test"]),
    ("pa_line_items", ["jc_43", "jc_62", "cube_43", "cube_61"]),
    ("sa_line_items", ["jc"]),
    ("packing_items", ["packing_type", "value"]),
    ("gcc_rows", ["color", "bb", "defect_reason", "note"]),
    ("gcc_summary", ["sb", "nc"]),
    ("loading_line_items", ["product_type", "exp", "dom"]),
]


def _jsonable(v):
    """Decimal والتواريخ مش بتتحول لـJSON تلقائياً."""
    import datetime as _dt

    if isinstance(v, Decimal) or isinstance(v, _dt.date):
        return str(v)
    return v


def snapshot_entry(entry):
    """نسخة كاملة من بيانات الإدخال الحالية (تتحفظ في سجل التعديل قبل أي تغيير)."""
    data = {
        "general_notes": entry.general_notes,
        "unit": entry.unit,
        "entry_date": str(entry.entry_date),
        "shift": entry.shift,
    }
    for name, fields in _SNAPSHOT_RELATED:
        if name in ("sop_line_items", "loading_line_items", "gcc_rows"):
            data[name] = [
                {f: _jsonable(getattr(obj, f)) for f in fields}
                for obj in getattr(entry, name).all()
            ]
        elif name == "packing_items":
            data[name] = [
                {
                    "factory": li.packing_type.factory,
                    "type": li.packing_type.name,
                    "value": _jsonable(li.value),
                }
                for li in entry.packing_items.select_related("packing_type")
            ]
        else:
            obj = getattr(entry, name, None)
            data[name] = (
                {f: _jsonable(getattr(obj, f)) for f in fields} if obj else None
            )
    lines = []
    for l in entry.dcp_reason_lines.select_related("reason"):
        lines.append(
            {"category": l.reason.category, "reason": l.reason.name, "qty": l.qty}
        )
    data["dcp_reason_lines"] = lines
    return data

# تسلسل الورديات المتكرر خلال اليوم: أول إدخال = الأولى، تاني = الثانية، تالت = الثالثة
SHIFT_SEQUENCE = [Shift.SHIFT_1, Shift.SHIFT_2, Shift.SHIFT_3]


def next_shift_for(unit, entry_date):
    """يرجّع أول وردية متاحة بالترتيب (1 ثم 2 ثم 3) لوحد معين في يوم معين،
    أو None لو الورديات الثلاثة اتسجلت بالفعل."""
    taken = set(
        ShiftEntry.objects.filter(unit=unit, entry_date=entry_date).values_list(
            "shift", flat=True
        )
    )
    for s in SHIFT_SEQUENCE:
        if s not in taken:
            return s
    return None


def header_context(request, unit, entry=None):
    """بيانات الهيدر الجاهزة تلقائياً: المستخدم، التاريخ، الوحدة، الوردية."""
    if entry is not None:
        return {
            "header_user": entry.submitted_by.get_full_name()
            or entry.submitted_by.username,
            "header_date": entry.entry_date,
            "header_shift": entry.get_shift_display(),
        }
    return {
        "header_user": request.user.get_full_name() or request.user.username,
        "header_date": timezone.localdate(),
        "header_shift_label": dict(Shift.choices)[next_shift_for(unit, timezone.localdate())]
        if next_shift_for(unit, timezone.localdate())
        else None,
    }

SOP_UNIT_TEMPLATES = {
    "SOP_A": "entry_sop.html",
    "SOP_B": "entry_sop.html",
    "SOP_C": "entry_sop.html",
    "SOP_D": "entry_sop.html",
    "G_SOP": "entry_sop.html",
    "C_PACKING": "entry_sop.html",
}


def unit_label(unit):
    return dict(ShiftEntry._meta.get_field("unit").choices)[unit]


def get_unit_or_404(unit):
    if unit not in Unit.values:
        raise Http404
    return unit


def user_can_access_unit(user, unit):
    return user.is_manager or user.assigned_unit == unit


def _first(queryset):
    return queryset.first()


def build_sop_rows(unit, data=None, entry=None):
    rows = []
    for pt in SOP_PRODUCT_TYPES[unit]:
        instance = None
        if entry is not None:
            instance = entry.sop_line_items.filter(product_type=pt).first()
        form = SOPLineItemForm(data, instance=instance, prefix=f"row_{pt}")
        show_car = pt == ProductType.BULK
        rows.append({"form": form, "label": ProductType(pt).label, "show_car": show_car})
    return rows


def save_sop_rows(entry, data):
    for pt in SOP_PRODUCT_TYPES[entry.unit]:
        form = SOPLineItemForm(data, prefix=f"row_{pt}")
        if form.is_valid():
            obj = form.save(commit=False)
            obj.shift_entry = entry
            obj.product_type = pt
            obj.save()


def build_loading_rows(data=None, entry=None):
    rows = []
    for pt in LoadingProductType:
        instance = None
        if entry is not None:
            instance = entry.loading_line_items.filter(product_type=pt).first()
        form = LoadingLineItemForm(data, instance=instance, prefix=f"row_{pt}")
        rows.append({"form": form, "label": LoadingProductType(pt).label})
    return rows


def save_loading_rows(entry, data):
    for pt in LoadingProductType:
        form = LoadingLineItemForm(data, prefix=f"row_{pt}")
        if form.is_valid():
            obj = form.save(commit=False)
            obj.shift_entry = entry
            obj.product_type = pt
            obj.save()


def dcp_reason_context(entry, category):
    """خيارات القائمة المنسدلة + الصفوف المحفوظة (عدد مفتوح)."""
    qs = DCPReason.objects.filter(category=category, is_active=True).order_by(
        "order", "id"
    )
    saved = []
    if entry is not None:
        saved = list(
            entry.dcp_reason_lines.filter(reason__category=category).order_by("id")
        )
    if saved:
        pairs = [{"reason_id": l.reason_id, "qty": l.qty} for l in saved]
    else:
        pairs = [{"reason_id": None, "qty": None}]
    return {"options": qs, "pairs": pairs}


def build_dcp_forms(data=None, entry=None):
    def single(form_class, related_name, prefix):
        instance = getattr(entry, related_name, None) if entry is not None else None
        return form_class(data, instance=instance, prefix=prefix)

    return {
        "bb": single(DCPBBRowForm, "dcp_bb", "bb"),
        "sb": single(DCPSBRowForm, "dcp_sb", "sb"),
        "as_row": single(DCPASRowForm, "dcp_as", "as"),
        "tests": single(DCPTestsForm, "dcp_tests", "tests"),
        "white": dcp_reason_context(entry, DCPReasonCategory.WHITE),
        "nc": dcp_reason_context(entry, DCPReasonCategory.NC),
    }


def save_dcp_forms(entry, data):
    """حفظ جداول DCP مع التحقق: مجموع أسباب الأبيض = White ومجموع الأحمر = Red."""
    from django.core.exceptions import ValidationError

    def single(form_class, related_name, prefix):
        form = form_class(data, instance=getattr(entry, related_name, None), prefix=prefix)
        if not form.is_valid():
            return None
        obj = form.save(commit=False)
        obj.shift_entry = entry
        obj.save()
        return obj

    bb = single(DCPBBRowForm, "dcp_bb", "bb")
    sb = single(DCPSBRowForm, "dcp_sb", "sb")
    as_row = single(DCPASRowForm, "dcp_as", "as")
    tests = single(DCPTestsForm, "dcp_tests", "tests")
    if bb is None or sb is None or as_row is None or tests is None:
        raise ValidationError("بيانات DCP غير مكتملة.")

    def collect_lines(prefix, category):
        reasons = data.getlist(f"{prefix}_reason")
        qtys = data.getlist(f"{prefix}_qty")
        lines, total = [], 0
        for raw_r, raw_q in zip(reasons, qtys):
            if not raw_r:
                continue
            try:
                qty = int(raw_q or 0)
            except ValueError:
                qty = 0
            reason = DCPReason.objects.filter(pk=raw_r, category=category).first()
            if reason and qty > 0:
                lines.append(DCPReasonLine(shift_entry=entry, reason=reason, qty=qty))
                total += qty
        return lines, total

    white_lines, white_total = collect_lines("w", DCPReasonCategory.WHITE)
    nc_lines, nc_total = collect_lines("n", DCPReasonCategory.NC)

    # منع تكرار نفس السبب مرتين في نفس الجدول
    w_ids = [l.reason_id for l in white_lines]
    n_ids = [l.reason_id for l in nc_lines]
    if len(w_ids) != len(set(w_ids)) or len(n_ids) != len(set(n_ids)):
        raise ValidationError("لا يمكن اختيار نفس السبب أكثر من مرة في نفس الجدول.")

    entry.dcp_reason_lines.all().delete()
    DCPReasonLine.objects.bulk_create(white_lines + nc_lines)

    if white_total != bb.white:
        raise ValidationError(
            f"مجموع أسباب الأبيض ({white_total}) لا يساوي عدد الأبيض في جدول B.B ({bb.white})."
        )
    if nc_total != bb.red:
        raise ValidationError(
            f"مجموع أسباب الأحمر ({nc_total}) لا يساوي عدد الأحمر في جدول B.B ({bb.red})."
        )


def all_row_forms_valid(forms_iterable):
    return all(f.is_valid() for f in forms_iterable)


@login_required
def dashboard(request):
    today = timezone.localdate()
    context = {"today": today}
    if request.user.is_manager:
        context["units"] = [
            {"value": u, "label": label}
            for u, label in ShiftEntry._meta.get_field("unit").choices
        ]
        context["is_manager"] = True
    else:
        unit = request.user.assigned_unit
        context["unit"] = unit
        context["unit_label"] = unit_label(unit)
        context["entries_today"] = ShiftEntry.objects.filter(unit=unit, entry_date=today)
        context["is_manager"] = False
    return render(request, "dashboard.html", context)


def _entry_template(unit):
    if unit in SOP_UNIT_TEMPLATES:
        return SOP_UNIT_TEMPLATES[unit]
    templates = {
        "DCP": "entry_dcp.html",
        "PA": "entry_packing.html",
        "SA": "entry_packing.html",
        "GCC1": "entry_gcc.html",
        "GCC2": "entry_gcc.html",
        "LOADING": "entry_loading.html",
    }
    return templates[unit]


def packing_sides(entry_date=None, shift=None, entry=None):
    """قيم عرض جدولي PA/SA معاً: الأنواع + القيم المحفوظة (للعرض فقط)."""
    sides = {}
    for factory in ("PA", "SA"):
        types = list(
            PackingType.objects.filter(factory=factory, is_active=True).order_by(
                "order", "id"
            )
        )
        src = entry
        if src is None and entry_date and shift:
            src = ShiftEntry.objects.filter(
                unit=factory, entry_date=entry_date, shift=shift
            ).first()
        values = (
            {li.packing_type_id: li.value for li in src.packing_items.all()}
            if src
            else {}
        )
        total = sum((v for v in values.values()), ZERO)
        sides[factory] = {
            "types": types,
            "values": values,
            "total": total,
            "label": dict(PackingFactory.choices)[factory],
        }
    return sides


def _build_all_forms(unit, data=None, entry=None):
    ctx = {
        "shift_form": ShiftEntryForm(data, instance=entry),
        "rows": [],
        "dcp": None,
        "pa_form": None,
        "sa_form": None,
        "pa_table": None,
        "sa_table": None,
        "bulk_form": None,
    }
    if unit in SOP_UNIT_TEMPLATES:
        ctx["rows"] = build_sop_rows(unit, data, entry)
        if unit in SOP_BULK_UNITS:
            instance = getattr(entry, "bulk_log", None) if entry is not None else None
            ctx["bulk_form"] = BulkLogForm(data, instance=instance, prefix="bulk")
    elif unit in ("GCC1", "GCC2"):
        ctx["gcc_color_forms"] = [
            {
                "color": c.value,
                "label": c.name,
                "form": gcc_color_form(c.value, data=data, entry=entry),
            }
            for c in GCCColor
        ]
        summary = getattr(entry, "gcc_summary", None) if entry is not None else None
        ctx["gcc_sum_form"] = GCCSummaryForm(data, instance=summary, prefix="gcc_sum")
    elif unit == "LOADING":
        ctx["rows"] = build_loading_rows(data, entry)
    elif unit == "DCP":
        ctx["dcp"] = build_dcp_forms(data, entry)
    elif unit in ("PA", "SA"):
        ctx["pa_table"] = PackingTableForm(
            PackingFactory.PA, data=data,
            entry=entry if entry is not None else None, prefix="pa",
        )
        ctx["sa_table"] = PackingTableForm(
            PackingFactory.SA, data=data,
            entry=entry if entry is not None else None, prefix="sa",
        )
    return ctx


def _flatten_forms(ctx):
    forms = [ctx["shift_form"]]
    forms.extend(r["form"] for r in ctx["rows"])
    if ctx.get("gcc_color_forms"):
        forms.extend(r["form"] for r in ctx["gcc_color_forms"])
        if ctx.get("gcc_sum_form"):
            forms.append(ctx["gcc_sum_form"])
    if ctx.get("bulk_form"):
        forms.append(ctx["bulk_form"])
    if ctx["dcp"]:
        d = ctx["dcp"]
        forms.append(d["bb"])
        forms.append(d["sb"])
        forms.append(d["as_row"])
        forms.append(d["tests"])
    if ctx["pa_form"]:
        forms.append(ctx["pa_form"])
    if ctx["sa_form"]:
        forms.append(ctx["sa_form"])
    if ctx.get("pa_table"):
        forms.append(ctx["pa_table"])
    if ctx.get("sa_table"):
        forms.append(ctx["sa_table"])
    return forms


def _save_related(entry, data):
    unit = entry.unit
    if unit in SOP_UNIT_TEMPLATES:
        save_sop_rows(entry, data)
        if unit in SOP_BULK_UNITS:
            form = BulkLogForm(
                data, instance=getattr(entry, "bulk_log", None), prefix="bulk"
            )
            if form.is_valid():
                obj = form.save(commit=False)
                obj.shift_entry = entry
                obj.save()
    elif unit in ("GCC1", "GCC2"):
        for c in GCCColor:
            form = gcc_color_form(c.value, data=data, entry=entry)
            if not form.is_valid():
                raise ValidationError("بيانات GCC غير صحيحة.")
            obj = form.save(commit=False)
            obj.shift_entry = entry
            obj.color = c.value
            obj.save()
        summary = getattr(entry, "gcc_summary", None)
        sum_form = GCCSummaryForm(data, instance=summary, prefix="gcc_sum")
        if sum_form.is_valid():
            obj = sum_form.save(commit=False)
            obj.shift_entry = entry
            obj.save()
    elif unit == "LOADING":
        save_loading_rows(entry, data)
    elif unit == "DCP":
        save_dcp_forms(entry, data)
    elif unit in ("PA", "SA"):
        for factory, prefix in ((PackingFactory.PA, "pa"), (PackingFactory.SA, "sa")):
            table = PackingTableForm(factory, data=data, prefix=prefix)
            if not table.is_valid():
                raise ValidationError("بيانات %s غير صحيحة." % factory)
            cleaned = table.cleaned_data
            for t in table.types:
                value = cleaned.get(f"type_{t.pk}") or 0
                PackingLineItem.objects.update_or_create(
                    shift_entry=entry,
                    packing_type=t,
                    defaults={"value": value},
                )


@login_required
def entry_new(request, unit):
    get_unit_or_404(unit)
    if not user_can_access_unit(request.user, unit):
        messages.error(request, "لا تملك صلاحية إدخال بيانات لهذه الوحدة.")
        return redirect("dashboard")

    today = timezone.localdate()

    # الوردية بتتحدد أوتوماتيك بالتسلسل: لو الورديات الثلاثة خلصت نمنع إدخال جديد
    next_shift = next_shift_for(unit, today)
    if next_shift is None:
        messages.info(
            request,
            f"تم تسجيل ورديات اليوم الثلاثة لوحدة {unit_label(unit)} بالفعل. "
            "لتعديل بيانات استخدم زر تعديل من لوحة التحكم.",
        )
        return redirect("dashboard")

    if request.method == "POST":
        ctx = _build_all_forms(unit, data=request.POST)
        shift_form = ctx["shift_form"]
        if all_row_forms_valid(_flatten_forms(ctx)):
            try:
                with transaction.atomic():
                    entry = shift_form.save(commit=False)
                    entry.unit = unit
                    entry.submitted_by = request.user
                    # التاريخ والوردية يتحسموا على السيرفر مش من المستخدم
                    entry.entry_date = today
                    entry.shift = next_shift
                    entry.full_clean()
                    entry.save()
                    _save_related(entry, request.POST)
                return redirect("entry_done", unit=unit, pk=entry.pk)
            except IntegrityError:
                shift_form.add_error(
                    None,
                    "تم إدخال هذه الوردية بالفعل لهذا اليوم والوحدة. عدّل الإدخال الموجود.",
                )
                messages.warning(request, "هذا الإدخال موجود بالفعل.")
            except ValidationError as ex:
                msgs = list(getattr(ex, "messages", []) or [str(ex)])
                for m in msgs:
                    shift_form.add_error(None, m)
    else:
        ctx = _build_all_forms(unit)

    entries_today = ShiftEntry.objects.filter(unit=unit, entry_date=today)
    existing_shifts = {e.shift: e for e in entries_today}
    return render(
        request,
        _entry_template(unit),
        {
            "unit": unit,
            "unit_label": unit_label(unit),
            "today": today,
            "entries_today": entries_today,
            "existing_shifts": existing_shifts,
            "has_bulk": unit in SOP_BULK_UNITS,
            "pack_sides": packing_sides(today, next_shift) if unit in ("PA", "SA") else None,
            **header_context(request, unit),
            **ctx,
        },
    )


@login_required
def entry_edit(request, unit, entry_date, shift):
    get_unit_or_404(unit)
    entry = get_object_or_404(
        ShiftEntry, unit=unit, entry_date=entry_date, shift=shift
    )
    if not (request.user.is_manager or entry.submitted_by == request.user):
        messages.error(request, "لا تملك صلاحية تعديل هذا الإدخال.")
        return redirect("dashboard")

    if request.method == "POST":
        ctx = _build_all_forms(unit, data=request.POST, entry=entry)
        if all_row_forms_valid(_flatten_forms(ctx)):
            try:
                with transaction.atomic():
                    snapshot = snapshot_entry(entry)  # نسخة البيانات قبل التعديل
                    ctx["shift_form"].save()
                    _save_related(entry, request.POST)
                    EntryRevision.objects.create(
                        shift_entry=entry,
                        edited_by=request.user,
                        data=snapshot,
                    )
                messages.success(request, "تم تعديل الإدخال بنجاح.")
                return redirect("entry_done", unit=unit, pk=entry.pk)
            except IntegrityError:
                ctx["shift_form"].add_error(
                    None,
                    "يوجد إدخال آخر لنفس الوحدة والتاريخ والوردية.",
                )
            except ValidationError as ex:
                for m in list(getattr(ex, "messages", []) or [str(ex)]):
                    ctx["shift_form"].add_error(None, m)
    else:
        ctx = _build_all_forms(unit, entry=entry)

    return render(
        request,
        _entry_template(unit),
        {
            "unit": unit,
            "unit_label": unit_label(unit),
            "editing": True,
            "entry": entry,
            "today": timezone.localdate(),
            "has_bulk": unit in SOP_BULK_UNITS,
            "pack_sides": packing_sides(entry=entry) if unit in ("PA", "SA") else None,
            **header_context(request, unit, entry=entry),
            **ctx,
        },
    )


def entry_summary_context(entry):
    ctx = {}
    unit = entry.unit
    if unit in SOP_UNIT_TEMPLATES:
        items = entry.sop_line_items.all()
        ctx["sop_items"] = items
        ctx["sop_total"] = sum((i.total for i in items), ZERO)
    elif unit == "DCP":
        ctx["dcp_bb"] = getattr(entry, "dcp_bb", None)
        ctx["dcp_sb"] = getattr(entry, "dcp_sb", None)
        ctx["dcp_as"] = getattr(entry, "dcp_as", None)
        ctx["dcp_tests"] = getattr(entry, "dcp_tests", None)
        ctx["dcp_white_lines"] = entry.dcp_reason_lines.filter(
            reason__category="WHITE"
        ).select_related("reason")
        ctx["dcp_nc_lines"] = entry.dcp_reason_lines.filter(
            reason__category="NC"
        ).select_related("reason")
    elif unit in ("PA", "SA"):
        items = entry.packing_items.select_related("packing_type")
        ctx["pa_items"] = [i for i in items if i.packing_type.factory == "PA"]
        ctx["sa_items"] = [i for i in items if i.packing_type.factory == "SA"]
    elif unit in ("GCC1", "GCC2"):
        ctx["gcc_rows"] = entry.gcc_rows.all()
        summary = getattr(entry, "gcc_summary", None)
        ctx["gcc_summary"] = summary
        sb_val = summary.sb if summary else 0
        ctx["gcc_total"] = sum(r.bb for r in ctx["gcc_rows"]) + sb_val
    elif unit == "LOADING":
        ctx["loading_items"] = entry.loading_line_items.all()
    return ctx


@login_required
def entry_done(request, unit, pk):
    get_unit_or_404(unit)
    entry = get_object_or_404(ShiftEntry, pk=pk, unit=unit)
    can_edit = request.user.is_manager or entry.submitted_by == request.user
    return render(
        request,
        "confirm.html",
        {
            "entry": entry,
            "unit_label": unit_label(unit),
            "can_edit": can_edit,
            **entry_summary_context(entry),
        },
    )


def _period_or_redirect(date_from, date_to):
    """ترتيب الفترة: لو النهاية قبل البداية نعكسها."""
    if date_from and date_to and date_to < date_from:
        return date_to, date_from
    return date_from, date_to


@login_required
def daily_report(request, date_from=None, date_to=None):
    if not request.user.is_manager:
        messages.error(request, "التقرير اليومي متاح للمدير فقط.")
        return redirect("dashboard")
    date_from = date_from or timezone.localdate()
    date_from, date_to = _period_or_redirect(date_from, date_to)
    context = full_daily_report(date_from, date_to)
    context["today"] = timezone.localdate()
    # تنقّل يوم سابق/تالي يظهر فقط في تقرير اليوم الواحد
    if not date_to or date_to == date_from:
        context["prev_date"] = date_from - timezone.timedelta(days=1)
        context["next_date"] = date_from + timezone.timedelta(days=1)
    return render(request, "report.html", context)


@login_required
def daily_report_pdf(request, date_from=None, date_to=None):
    if not request.user.is_manager:
        messages.error(request, "التقرير اليومي متاح للمدير فقط.")
        return redirect("dashboard")
    date_from = date_from or timezone.localdate()
    date_from, date_to = _period_or_redirect(date_from, date_to)

    # الاستيراد هنا (lazy import) عشان WeasyPrint يحتاج مكتبات نظام (Cairo/Pango)
    # قد لا تكون متاحة في كل بيئة، فمنعزلها عن باقي التطبيق حتى لا يفشل التشغيل كله لو غابت.
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        messages.error(
            request,
            "خدمة PDF غير متاحة على هذه البيئة. ثبّت weasyprint ومكتبات النظام المطلوبة "
            "(Pango/Cairo) أو استخدم زر الطباعة من المتصفح.",
        )
        return redirect("report", date_from)

    context = full_daily_report(date_from, date_to)
    context["charts"] = generate_report_charts(context["charts"])
    context["doc_control"] = settings.QC_REPORT_DOC_CONTROL

    html = render_to_string("report_pdf.html", context, request=request)

    pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    if date_to and date_to != date_from:
        filename = f"QC_Report_{date_from.strftime('%d-%m-%Y')}_{date_to.strftime('%d-%m-%Y')}.pdf"
    else:
        filename = f"QC_Report_{date_from.strftime('%d-%m-%Y')}.pdf"
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


_REVISION_TITLES = [
    ("sop_line_items", "بنود SOP"),
    ("bulk_log", "عربيات BULK"),
    ("dcp_bb", "جدول B.B"),
    ("dcp_sb", "جدول S.B"),
    ("dcp_as", "جدول S.B AS"),
    ("dcp_tests", "الاختبارات"),
    ("pa_line_items", "بنود PA (قديم)"),
    ("sa_line_items", "بنود SA (قديم)"),
    ("packing_items", "بنود التعبئة (PA/SA)"),
    ("gcc_rows", "جدول GCC - الألوان"),
    ("gcc_summary", "SB / NC"),
    ("loading_line_items", "بنود التحميل"),
    ("dcp_reason_lines", "أسباب الأبيض والأحمر"),
]


def _revision_sections(data):
    """يحوّل نسخة JSON لجداول جاهزة للعرض: عنوان + رؤوس أعمدة + صفوف."""
    sections = []
    for key, title in _REVISION_TITLES:
        rows = data.get(key)
        if not rows:
            continue
        if isinstance(rows, dict):
            rows = [rows]
        header = [k.replace("_", " ").title() for k in rows[0]]
        body = [[str(v) for v in r.values()] for r in rows]
        sections.append({"title": title, "header": header, "rows": body})
    return sections


@login_required
def entry_history(request, unit, pk):
    get_unit_or_404(unit)
    entry = get_object_or_404(ShiftEntry, pk=pk, unit=unit)
    if not (request.user.is_manager or entry.submitted_by == request.user):
        messages.error(request, "لا تملك صلاحية عرض سجل التعديلات.")
        return redirect("dashboard")
    history = [
        {
            "rev": rev,
            "notes": rev.data.get("general_notes"),
            "sections": _revision_sections(rev.data),
        }
        for rev in entry.revisions.select_related("edited_by")
    ]
    return render(
        request,
        "entry_history.html",
        {"entry": entry, "unit_label": unit_label(unit), "history": history},
    )


@login_required
def shift_report_picker(request):
    """صفحة اختيار اليوم والوردية لتقرير الوردية لكل المصانع."""
    if not request.user.is_manager:
        messages.error(request, "تقارير متاحة للمدير فقط.")
        return redirect("dashboard")
    report_date = request.GET.get("date")
    shift = request.GET.get("shift")
    parsed = parse_date(report_date) if report_date else None
    if parsed and shift in Shift.values:
        return redirect("shift_report", parsed, shift)
    return render(request, "shift_report_picker.html", {"today": timezone.localdate(), "shifts": Shift.choices})


@login_required
def shift_report(request, report_date, shift):
    if not request.user.is_manager:
        messages.error(request, "التقارير متاحة للمدير فقط.")
        return redirect("dashboard")
    if shift not in Shift.values:
        raise Http404
    context = full_shift_report(report_date, shift)
    context["today"] = timezone.localdate()
    context["prev_date"] = report_date - timezone.timedelta(days=1)
    context["next_date"] = report_date + timezone.timedelta(days=1)
    return render(request, "report_shift.html", context)


def _format_field_value(field, value):
    """تنسيق قيمة حقل للعرض في الجداول: اختيارات بالتسمية، أرقام بدون أصفار زائدة."""
    if value is None or value == "":
        return "—"
    if getattr(field, "choices", None):
        return dict(field.choices).get(value, str(value))
    if isinstance(value, Decimal):
        return f"{value:f}".rstrip("0").rstrip(".") if "." in f"{value:f}" else f"{value:f}"
    return str(value)


def entry_tables(entry):
    """يبني جداول كل بيانات الإدخال تلقائياً من علاقات الموديل،
    فأي حقل جديد يتضافر لموديل يظهر هنا من غير تعديل القالب."""
    sections = []
    for rel in ShiftEntry._meta.related_objects:
        accessor = rel.get_accessor_name()
        if rel.one_to_one:
            obj = getattr(entry, accessor, None)
            instances = [obj] if obj else []
        else:
            instances = list(getattr(entry, accessor).all())
        if not instances:
            continue
        model = rel.related_model
        fields = [
            f
            for f in model._meta.concrete_fields
            if not f.primary_key and f.name != "shift_entry"
        ]
        section = {
            "title": MODEL_TITLES.get(
                model.__name__,
                model._meta.verbose_name_plural
                or model._meta.verbose_name
                or model.__name__,
            ),
            "headers": [str(f.verbose_name) for f in fields],
            "rows": [
                [_format_field_value(f, getattr(inst, f.name)) for f in fields]
                for inst in instances
            ],
        }
        if len(instances) == 1:
            # لجداول الصف الواحد نجهز أزواج (اسم: قيمة) للعرض المدمج
            row = section["rows"][0]
            section["pairs"] = [
                {"label": h, "value": v} for h, v in zip(section["headers"], row)
            ]
        sections.append(section)
    return sections


def _filtered_entries(request):
    """الإدخالات بعد تطبيق صلاحيات المستخدم وفلاتر الوحدة/الفترة."""
    qs = ShiftEntry.objects.select_related("submitted_by")
    if not request.user.is_manager:
        qs = qs.filter(unit=request.user.assigned_unit)
    unit = request.GET.get("unit") or ""
    date_from = request.GET.get("from") or ""
    date_to = request.GET.get("to") or ""
    if unit:
        qs = qs.filter(unit=unit)
    try:
        if date_from:
            qs = qs.filter(entry_date__gte=date_from)
        if date_to:
            qs = qs.filter(entry_date__lte=date_to)
    except (ValueError, ValidationError):
        date_from = date_to = ""
    return qs.order_by("-entry_date", "shift", "-pk"), date_from, date_to


@login_required
def records_list(request):
    qs, date_from, date_to = _filtered_entries(request)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    params = request.GET.copy()
    params.pop("page", None)
    querystring = params.urlencode()

    if request.user.is_manager:
        available_units = Unit.choices
    else:
        available_units = [(request.user.assigned_unit, unit_label(request.user.assigned_unit))]

    return render(
        request,
        "records.html",
        {
            "page_obj": page_obj,
            "units": available_units,
            "sel_unit": request.GET.get("unit") or "",
            "date_from": date_from,
            "date_to": date_to,
            "querystring": querystring,
            "total": paginator.count,
        },
    )


@login_required
def record_detail(request, pk):
    entry = get_object_or_404(
        ShiftEntry.objects.select_related("submitted_by"), pk=pk
    )
    if not (request.user.is_manager or entry.submitted_by == request.user
            or entry.unit == request.user.assigned_unit):
        messages.error(request, "لا تملك صلاحية عرض هذا الإدخال.")
        return redirect("dashboard")
    can_edit = request.user.is_manager or entry.submitted_by == request.user
    return render(
        request,
        "record_detail.html",
        {
            "entry": entry,
            "unit_label": unit_label(entry.unit),
            "can_edit": can_edit,
            "sections": entry_tables(entry),
        },
    )


def _excel_cell(field, value):
    """قيمة حقل كقيمة إكسيل أصلية (تاريخ/رقم) بدل نص."""
    if value is None or value == "":
        return ""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bool):
        return "نعم" if value else "لا"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M") if value.hour or value.minute else value.strftime("%Y-%m-%d")
    if getattr(field, "choices", None):
        return dict(field.choices).get(value, str(value))
    return value


def _style_sheet(ws, ncols, freeze=True):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    if freeze:
        ws.freeze_panes = "A2"
    # عرض أعمدة معقول بناء على أطول قيمة في أول 200 صف
    for c in range(1, ncols + 1):
        width = 12
        for r in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                width = max(width, min(len(str(v)) + 4, 40))
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = width


@login_required
def records_export(request):
    from openpyxl import Workbook

    qs, date_from, date_to = _filtered_entries(request)

    wb = Workbook()

    # ورقة السجل العام: صف لكل إدخال وردية
    main = wb.active
    main.title = "السجل"
    main.sheet_view.rightToLeft = True
    main.append(["التاريخ", "الوردية", "الوحدة", "الموظف", "ملاحظات"])
    for e in qs:
        main.append(
            [
                e.entry_date,
                e.get_shift_display(),
                e.get_unit_display(),
                e.submitted_by.get_full_name() or e.submitted_by.username,
                e.general_notes or "",
            ]
        )
        main.cell(row=main.max_row, column=1).number_format = "yyyy-mm-dd"
    _style_sheet(main, 5)

    # ورقة لكل (وحدة × جدول بيانات) فيها بيانات داخل الفترة
    unit_codes = [u for u, _ in Unit.choices]
    units_present = set(qs.values_list("unit", flat=True))
    for code in unit_codes:
        if code not in units_present:
            continue
        for rel in ShiftEntry._meta.related_objects:
            model = rel.related_model
            rows_qs = (
                model.objects.filter(shift_entry__in=qs, shift_entry__unit=code)
                .select_related("shift_entry")
                .order_by("-shift_entry__entry_date", "-shift_entry__pk")
            )
            fields = [
                f
                for f in model._meta.concrete_fields
                if not f.primary_key and f.name != "shift_entry"
            ]
            headers = ["التاريخ", "الوردية"] + [str(f.verbose_name) for f in fields]
            data_rows = []
            for inst in rows_qs:
                e = inst.shift_entry
                data_rows.append(
                    [e.entry_date, e.get_shift_display()]
                    + [
                        _excel_cell(f, getattr(inst, f.name))
                        for f in fields
                    ]
                )
            if not data_rows:
                continue

            label = MODEL_TITLES.get(
                model.__name__,
                str(
                    model._meta.verbose_name_plural
                    or model._meta.verbose_name
                    or model.__name__
                ),
            )
            ws = wb.create_sheet(title=f"{code} - {label}"[:31])
            ws.sheet_view.rightToLeft = True
            ws.append(headers)
            for row in data_rows:
                ws.append(row)
                ws.cell(row=ws.max_row, column=1).number_format = "yyyy-mm-dd"
            _style_sheet(ws, len(headers))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    parts = [p.replace("-", "") or "الكل" for p in (date_from, date_to)]
    filename = f"QC_Records_{parts[0]}_{parts[1]}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
